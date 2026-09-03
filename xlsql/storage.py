import os
import threading

from openpyxl import Workbook, load_workbook

ID_COLUMN = "id"
META_SHEET = "__xlsql_meta__"


class XlsxError(Exception):
    pass


class Table:
    """One sheet in the database workbook == one table."""

    def __init__(self, wb, db, name, columns):
        self._wb = wb
        self._db = db
        self.name = name
        self.columns = columns
        self._lock = threading.Lock()

    @classmethod
    def load(cls, wb, db, name):
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise XlsxError(f'table "{name}" is empty')
        columns = [str(c) for c in rows[0]]
        data = [list(r) for r in rows[1:] if any(c is not None for c in r)]
        return cls(wb, db, name, columns), data

    @classmethod
    def create(cls, wb, db, name, columns):
        if ID_COLUMN not in columns:
            columns = [ID_COLUMN] + columns
        ws = wb.create_sheet(title=name)
        ws.append(columns)
        return cls(wb, db, name, columns)

    def read_all(self):
        """Return a list of row dicts."""
        with self._lock:
            _, data = Table.load(self._wb, self._db, self.name)
        rows = []
        for row in data:
            d = {}
            for i, col in enumerate(self.columns):
                d[col] = row[i] if i < len(row) else None
            rows.append(d)
        return rows

    def append(self, values):
        """Append one row. values is a dict column->value. Returns new id."""
        with self._lock:
            _, data = Table.load(self._wb, self._db, self.name)
            if data:
                maxid = max(int(r[0]) for r in data if r[0] is not None)
            else:
                maxid = 0
            newid = maxid + 1
            row = [newid]
            for col in self.columns[1:]:
                row.append(values.get(col))
            ws = self._wb[self.name]
            ws.append(row)
            self._db.save()
        return newid

    def write_all(self, rows):
        """rows is a list of row dicts. Rewrites the whole sheet."""
        with self._lock:
            ws = self._wb[self.name]
            ws.delete_rows(1, ws.max_row)
            ws.append(self.columns)
            for d in rows:
                ws.append([d.get(col) for col in self.columns])
            self._db.save()

    def drop(self):
        with self._lock:
            del self._wb[self.name]


class Database:
    """Manages a single .xlsx workbook where each sheet is a table."""

    def __init__(self, filepath):
        if not filepath.lower().endswith(".xlsx"):
            filepath = f"{filepath}.xlsx"
        self.filepath = filepath
        self._lock = threading.RLock()
        self._wb = self._open(filepath)

    @staticmethod
    def _ensure_parent(filepath):
        parent = os.path.dirname(os.path.abspath(filepath))
        if parent and not os.path.isdir(parent):
            os.makedirs(parent, exist_ok=True)

    @staticmethod
    def _open(filepath):
        if filepath and os.path.exists(filepath):
            return load_workbook(filepath)
        wb = Workbook()
        wb.remove(wb.active)
        return wb

    def save(self):
        with self._lock:
            self._ensure_parent(self.filepath)
            if not self._wb.sheetnames:
                self._wb.create_sheet(title=META_SHEET)
            self._wb.save(self.filepath)

    def _drop_meta(self):
        if self._wb.sheetnames == [META_SHEET]:
            del self._wb[META_SHEET]

    def _assert_name(self, name):
        if not self._valid_name(name):
            raise XlsxError(f"invalid table name: {name!r}")

    @staticmethod
    def _valid_name(name):
        return bool(name) and all(
            c.isalnum() or c == "_" for c in name
        ) and not name.startswith(".")

    def list_tables(self):
        with self._lock:
            return sorted(s for s in self._wb.sheetnames if s != META_SHEET)

    def table_exists(self, name):
        with self._lock:
            return name in self._wb.sheetnames

    def create_table(self, name, columns):
        with self._lock:
            self._assert_name(name)
            if name in self._wb.sheetnames:
                raise XlsxError(f'table "{name}" already exists')
            self._drop_meta()
            tbl = Table.create(self._wb, self, name, columns)
            self.save()
        return Table(self._wb, self, name, [ID_COLUMN] + columns)

    def get_table(self, name):
        with self._lock:
            self._assert_name(name)
            if name not in self._wb.sheetnames:
                raise XlsxError(f'table "{name}" does not exist')
            tbl, _ = Table.load(self._wb, self, name)
        return tbl

    def drop_table(self, name):
        with self._lock:
            tbl = self.get_table(name)
            tbl.drop()
            self.save()
